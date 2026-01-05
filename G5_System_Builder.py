import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import date

# --- PALETA DE CORES PREMIUM ---
C_HEADER_DARK = "000000"     # Preto Profundo para Headers
C_HEADER_ACCENT = "FFD700"   # Dourado para detalhes
C_BG_MAIN = "FFFFFF"         # Branco Fundo
C_BG_ALT = "F8F9FA"          # Cinza Bem Claro para alternância
C_TEXT_MAIN = "212529"       # Cinza Escuro Texto
C_BORDER = "DEE2E6"          # Borda Suave

# Cores de Flags/Tags
C_FLAG_BAD = "FFCDD2"        # Vermelho Claro (Nota 1/Ruim)
C_FLAG_AVG = "FFF9C4"        # Amarelo Claro (Nota 2/Media)
C_FLAG_GOOD = "C8E6C9"       # Verde Claro (Nota 3/Bom)
C_FLAG_INFO = "BBDEFB"       # Azul (Amistoso/Folga)

# Estilos Base
font_header = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
font_sub = Font(name='Segoe UI', size=10, bold=True, color="FFFFFF")
font_body = Font(name='Segoe UI', size=10, color=C_TEXT_MAIN)
font_body_bold = Font(name='Segoe UI', size=10, bold=True, color=C_TEXT_MAIN)

border_thin = Border(left=Side(style='thin', color=C_BORDER), 
                     right=Side(style='thin', color=C_BORDER), 
                     top=Side(style='thin', color=C_BORDER), 
                     bottom=Side(style='thin', color=C_BORDER))

border_thick_outer = Border(outline=True, top=Side(style='thick'), bottom=Side(style='thick'))

fill_header = PatternFill(start_color=C_HEADER_DARK, end_color=C_HEADER_DARK, fill_type="solid")
fill_accent = PatternFill(start_color="2C3E50", fill_type="solid") # Azul Petroleo

wb = Workbook()
del wb['Sheet']

# ==========================================
# 0. FUNÇÕES AUXILIARES
# ==========================================
def format_cell(ws, row, col, value, font=font_body, fill=None, align=None, border=border_thin):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    return cell

def apply_cond_formatting(ws, range_str):
    # Regras Visuais para Flags
    # 1 ou Ruim -> Vermelho
    ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'OR(A1="1", A1="Ruim", A1="F")'], fill=PatternFill(bgColor=C_FLAG_BAD)))
    # 2 ou Media -> Amarelo
    ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'OR(A1="2", A1="Na Média")'], fill=PatternFill(bgColor=C_FLAG_AVG)))
    # 3 ou Bom -> Verde
    ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'OR(A1="3", A1="Bom", A1="Se Destacou")'], fill=PatternFill(bgColor=C_FLAG_GOOD)))
    # Outros -> Azul
    ws.conditional_formatting.add(range_str, FormulaRule(formula=[f'OR(A1="FO", A1="AM", A1="CH")'], fill=PatternFill(bgColor=C_FLAG_INFO)))

# ==========================================
# 1. ABA CONFIG
# ==========================================
ws_config = wb.create_sheet("CONFIG")
ws_config.sheet_view.showGridLines = False

format_cell(ws_config, 1, 1, "CONFIGURAÇÕES DO SISTEMA", font=Font(size=16, bold=True), border=None)

# Tabelas
headers = ["CATEGORIA", "ANOS (TXT)"]
for i, h in enumerate(headers):
    format_cell(ws_config, 3, 2+i, h, font=font_header, fill=fill_header, align=Alignment(horizontal='center'))

# Dados Categorias (Importante: Formato Texto para busca)
cats = [
    ["Sub 17", "2009, 2010"], ["Sub 15", "2011"], ["Sub 14", "2012"],
    ["Sub 13", "2013"], ["Sub 12", "2014"], ["Sub 11", "2015"],
    ["Sub 10", "2016"], ["Sub 9", "2017"]
]
for i, r in enumerate(cats):
    format_cell(ws_config, 4+i, 2, r[0])
    format_cell(ws_config, 4+i, 3, r[1])

# Meses
ws_config['E3'] = "MESES"
full_months = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
for i, m in enumerate(full_months):
    ws_config.cell(row=4+i, column=5, value=m)

# Legendas (Range nomeado se quisessemos, mas faremos direto)
ws_config['G3'] = "FLAGS ATLETA"
flags_atl = ["1", "2", "3", "F", "DM"]
for i, f in enumerate(flags_atl):
    ws_config.cell(row=4+i, column=7, value=f)

ws_config['H3'] = "FLAGS TREINO"
flags_tr = ["1", "2", "3", "FO", "AM", "CH"]
for i, f in enumerate(flags_tr):
    ws_config.cell(row=4+i, column=8, value=f)


# ==========================================
# 2. ABA DB_ATLETAS
# ==========================================
ws_db = wb.create_sheet("DB_ATLETAS")
format_cell(ws_db, 2, 2, "BANCO DE DADOS DE ATLETAS", font=Font(size=16, bold=True), border=None)

# Headers
headers_db = ["Nome Completo", "Data Nasc", "Ano (Auto)", "Categoria (Auto)", "Posição", "Status"]
for i, h in enumerate(headers_db):
    format_cell(ws_db, 4, 2+i, h, font=font_header, fill=fill_header, align=Alignment(horizontal='center'))

ws_db.column_dimensions['B'].width = 30
ws_db.column_dimensions['C'].width = 15

# Dados Dummy
dummy_data = [
    ["Gabriel Silva", "12/05/2012", "Goleiro"],
    ["Lucas Pereira", "08/02/2012", "Atacante"],
    ["Matheus Oliveira", "15/09/2011", "Zagueiro"],
    ["Enzo Santos", "20/01/2013", "Meia"],
    ["Pedro Rocha", "05/11/2012", "Volante"],
    ["João Souza", "18/07/2011", "Lateral"]
]

for i, d in enumerate(dummy_data):
    r = 5+i
    ws_db.cell(row=r, column=2, value=d[0])
    ws_db.cell(row=r, column=3, value=d[1])
    ws_db.cell(row=r, column=6, value=d[2]) 
    ws_db.cell(row=r, column=7, value="ATIVO") # Coluna G Status
    
    # Formula Ano: =IF(C5="","",TEXT(C5, "yyyy")) -> Mais seguro que year() para textos
    ws_db.cell(row=r, column=4, value=f'=IF(C{r}="","",TEXT(C{r}, "yyyy"))')
    # Formula Categoria (Visual apenas): 
    ws_db.cell(row=r, column=5, value="Auto")

# Tabela Oficial (Para referencias estruturadas se precisar)
# range_db = f"B4:G{5+len(dummy_data)}"


# ==========================================
# 3. ABA TREINOS (A Estrela do Show)
# ==========================================
ws_t = wb.create_sheet("TREINOS")
ws_t.sheet_view.showGridLines = False # Look clean

# --- CONTROLES DE TOPO ---
# Fundo escuro no topo
for c in range(1, 36):
    ws_t.cell(row=1, column=c).fill = fill_accent
    ws_t.cell(row=2, column=c).fill = fill_accent

ws_t['B1'] = "CONTROLE DE TREINOS - G5"
ws_t['B1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_t['B1'].fill = fill_accent

# Seletores
format_cell(ws_t, 1, 5, "MÊS:", font=font_sub, fill=fill_accent, align=Alignment(horizontal='right'))
c_mes = format_cell(ws_t, 1, 6, "Janeiro", font=Font(bold=True), fill=PatternFill(start_color="FFFFFF", fill_type="solid"), align=Alignment(horizontal='center'))
dv_mes = DataValidation(type="list", formula1="'CONFIG'!$E$4:$E$15")
ws_t.add_data_validation(dv_mes)
dv_mes.add(c_mes)

format_cell(ws_t, 1, 8, "CATEGORIA:", font=font_sub, fill=fill_accent, align=Alignment(horizontal='right'))
c_cat = format_cell(ws_t, 1, 9, "Sub 14", font=Font(bold=True), fill=PatternFill(start_color="FFFFFF", fill_type="solid"), align=Alignment(horizontal='center'))
dv_cat = DataValidation(type="list", formula1="'CONFIG'!$B$4:$B$11")
ws_t.add_data_validation(dv_cat)
dv_cat.add(c_cat)

format_cell(ws_t, 1, 11, "ANO:", font=font_sub, fill=fill_accent, align=Alignment(horizontal='right'))
ws_t.cell(row=1, column=12, value=2026).font = font_sub # Fixo por enquanto
ws_t.cell(row=1, column=12).fill = fill_accent

# --- AUXILIARES OCULTOS ---
# Precisamos achar os anos da categoria para o filtro
# AA1 = Anos Texto
ws_t['AA1'] = f'=VLOOKUP(I1, CONFIG!B4:C11, 2, 0)' 
# AA2 = Numero Mes
ws_t['AA2'] = f'=MATCH(F1, CONFIG!E4:E15, 0)'
ws_t.column_dimensions['AA'].hidden = True


# --- GRID DE DIAS E SEMANAS ---
# Header começa na linha 4
start_col = 3
days = 31

# Linha de Semanas (Visual) - Opcional, vamos focar nos Dias
# Linha Dia Semana (S, T, Q...)
for d in range(1, 32):
    col = start_col + d - 1
    # Data: DATE(2026, MesNum, d)
    # Dia Semana: TEXT(..., "ddd")
    
    # Formula Complexa para não quebrar em dias inexistentes (30 Fev) || IF(MONTH(DATE...)=MesNum, ... , "")
    date_f = f'DATE($L$1, $AA$2, {d})'
    check_month = f'MONTH({date_f})=$AA$2'
    
    # Letra (S, T, Q)
    f_ddd = f'=IF(ISERROR({date_f}), "", IF({check_month}, LEFT(UPPER(TEXT({date_f}, "ddd")), 1), ""))'
    # Numero (01, 02)
    f_num = f'=IF(ISERROR({date_f}), "", IF({check_month}, {d}, ""))'
    
    c_ddd = format_cell(ws_t, 4, col, f_ddd, font=Font(bold=True, color="FFFFFF"), fill=fill_header, align=Alignment(horizontal='center'))
    c_num = format_cell(ws_t, 5, col, f_num, font=Font(bold=True), fill=PatternFill(start_color="E0E0E0", fill_type="solid"), align=Alignment(horizontal='center'), border=border_thin)

    # Largura fina padrão check
    ws_t.column_dimensions[get_column_letter(col)].width = 3.5

# --- CHECKLIST DE TREINOS ---
# Lista de Tipos de Treino (Hardcoded para layout)
tr_types = ["Físico", "Técnico", "Tático", "Vídeo", "Amistoso", "Coletivo", "Palestra", "Academia"]
curr_row = 6

for t in tr_types:
    # Label Lado Esquerdo
    format_cell(ws_t, curr_row, 1, t.upper(), font=Font(size=9, bold=True, color="566573"), align=Alignment(horizontal='right'))
    
    # Grid de Checkboxes (Simulado com 'x')
    for d in range(1, 32):
        col = start_col + d - 1
        c = ws_t.cell(row=curr_row, column=col)
        c.border = border_thin
        c.alignment = Alignment(horizontal='center')
    
    curr_row += 1

# Validação 'x' na grid
last_tr_row = curr_row - 1
dv_x = DataValidation(type="list", formula1='"x"', allow_blank=True)
ws_t.add_data_validation(dv_x)
dv_x.add(f"C6:{get_column_letter(start_col+30)}{last_tr_row}")

# Divisor
curr_row += 1
ws_t.row_dimensions[curr_row].height = 5 # Spacer

# --- AVALIAÇÃO DO DIA ---
curr_row += 1
format_cell(ws_t, curr_row, 1, "AVALIAÇÃO DO TREINO", font=Font(bold=True, color="FFFFFF"), fill=fill_accent, align=Alignment(horizontal='right'))
ws_t.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)

for d in range(1, 32):
    col = start_col + d - 1
    c = ws_t.cell(row=curr_row, column=col)
    c.border = border_thin
    c.fill = PatternFill(start_color="D4E6F1", fill_type="solid") # Azulzinho
    c.alignment = Alignment(horizontal='center', vertical='center')

# Validação do Treino (1-3, etc)
dv_tr_flag = DataValidation(type="list", formula1="'CONFIG'!$H$4:$H$9")
ws_t.add_data_validation(dv_tr_flag)
dv_tr_flag.add(f"C{curr_row}:{get_column_letter(start_col+30)}{curr_row}")


# --- LISTA DEADLIFT (ATLETAS) ---
curr_row += 2
format_cell(ws_t, curr_row, 1, "ATLETA", font=font_header, fill=fill_header, align=Alignment(horizontal='center'))
format_cell(ws_t, curr_row, 2, "POS", font=font_header, fill=fill_header, align=Alignment(horizontal='center'))
ws_t.column_dimensions['A'].width = 25
ws_t.column_dimensions['B'].width = 10

# Formula FILTER Robusta
# =FILTER(DB_ATLETAS!B:B, ISNUMBER(SEARCH(DB_ATLETAS!D:D, AA1)), "-")
# DB Col B = Nome
# DB Col D = Ano (Texto)
# AA1 = Lista de Anos permitidos "2012, 2013"
# SEARCH vai procurar o ano do atleta na lista permitida.

start_atl_row = curr_row + 1

# Coluna A (Nome)
f_atl_filter = f'=SORT(FILTER(DB_ATLETAS!B5:B100, ISNUMBER(SEARCH(DB_ATLETAS!D5:D100, $AA$1)), "Sem atletas nessa cat."))'
ws_t.cell(row=start_atl_row, column=1, value=f_atl_filter).font = font_body
ws_t.cell(row=start_atl_row, column=1).alignment = Alignment(indent=1)

# Coluna B (Posição) - PROCV Dinâmico no # (SPILL)
# =IF(A15#="","", VLOOKUP(A15#, DB_ATLETAS!B:E, 4, 0))
# Mas PROCV não faz spill nativo fácil sem LAMBDA ou MAP em versoes antigas.
# Vamos assumir que o usuário arrasta ou usamos uma formula simples apenas na primeira célula se ele tiver 365.
# Para garantir, vamos preencher 50 linhas com IFERROR(VLOOKUP) normal, pois FILTER despeja (Spill) mas as colunas ao lado não acompanham automaticamente sem Map.
# MELHOR: Usar LOOKUP Row a Row.
# Mas como o A despeja dinamicamente, não podemos prever onde acaba.
# TRUQUE: Deixar a coluna A despejar. Nas colunas de Grid, deixar formatado 50 linhas.

for r in range(start_atl_row, start_atl_row + 40):
    # Formatação Linha Atleta
    # Visual Zebrado check
    fill_row = C_BG_ALT if r % 2 == 0 else C_BG_MAIN
    fill_obj = PatternFill(start_color=fill_row, fill_type="solid")
    
    # Celula Nome e Pos (Preenchimento manual se Spill falhar?)
    # Não, Spill cuida dos dados. Nós cuidamos do visual.
    ws_t.cell(row=r, column=1).fill = fill_obj
    ws_t.cell(row=r, column=1).border = border_thin
    
    # Pos (Procv manual linha a linha pq é mais seguro q PROCV matricial complexo)
    # Se A{r} tiver valor, busca.
    ws_t.cell(row=r, column=2, value=f'=IF(A{r}="","",IFERROR(VLOOKUP(A{r},DB_ATLETAS!B:F,5,0),""))')
    ws_t.cell(row=r, column=2).fill = fill_obj
    ws_t.cell(row=r, column=2).border = border_thin
    ws_t.cell(row=r, column=2).alignment = Alignment(horizontal='center')
    
    # Grid de Notas
    for d in range(1, 32):
        col = start_col + d - 1
        c = ws_t.cell(row=r, column=col)
        c.border = border_thin
        c.fill = fill_obj
        c.alignment = Alignment(horizontal='center')

# Validação Atletas
dv_atl_flag = DataValidation(type="list", formula1="'CONFIG'!$G$4:$G$8")
ws_t.add_data_validation(dv_atl_flag)
dv_atl_flag.add(f"C{start_atl_row}:{get_column_letter(start_col+30)}{start_atl_row+39}")

# Conditional Formatting nas Flags (Visual)
# Aplicar em todo o range de notas de atletas e treinos
range_cond_atl = f"C{start_atl_row}:{get_column_letter(start_col+30)}{start_atl_row+39}"
range_cond_tr = f"C{start_atl_row-3}:{get_column_letter(start_col+30)}{start_atl_row-3}" # Linha avaliação treino

# Precisamos usar regras nativas do Excel em ingles nas fórmulas (As vezes conflitante com local).
# Mas openpyxl escreve fórmulas como strings.
# Usar cores de fundo baseadas no valor.

# Criando regras condicionais manuais (já definidas no helper acima, mas precisamos chamar)
# Como openpyxl conditional formatting é chato com formulas relativas 'A1', precisamos focar no range.
# O helper apply_cond_formatting usa A1 relativo. Excel ajusta.
apply_cond_formatting(ws_t, range_cond_atl)
apply_cond_formatting(ws_t, range_cond_tr)


# --- CONGELAR PAINÉIS ---
ws_t.freeze_panes = ws_t['C6']


# Salvar
file_name = "Sistema_G5_Futebol_Premium.xlsx"
wb.save(file_name)
print(f"Arquivo '{file_name}' gerado.")
